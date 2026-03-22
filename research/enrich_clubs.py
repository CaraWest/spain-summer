"""
RFEN Swim Club Enrichment Script
Filters clubs by name, then enriches with Google Places API data.
Output: spain-summer-2027/research/rfen_swim_clubs_enriched.xlsx
"""

import os
import sys
import time
import re
import requests
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv

# ── Config ────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
INPUT_FILE = BASE_DIR / "rfen_swim_clubs.xlsx"
OUTPUT_FILE = BASE_DIR / "rfen_swim_clubs_enriched.xlsx"

load_dotenv(BASE_DIR.parent / ".env.local")
API_KEY = os.getenv("GOOGLE_PLACES_API_KEY")

PLACES_URL = "https://places.googleapis.com/v1/places:searchText"
FIELD_MASK = ",".join([
    "places.displayName",
    "places.formattedAddress",
    "places.location",
    "places.websiteUri",
    "places.nationalPhoneNumber",
    "places.regularOpeningHours",
    "places.primaryType",
    "places.types",
])

REQUEST_DELAY = 0.2  # seconds between API calls


def places_search_text(query: str) -> requests.Response:
    return requests.post(
        PLACES_URL,
        headers={
            "Content-Type": "application/json",
            "X-Goog-Api-Key": API_KEY,
            "X-Goog-FieldMask": FIELD_MASK,
        },
        json={"textQuery": query, "languageCode": "es"},
        timeout=10,
    )


def exit_if_places_api_blocked(resp: requests.Response, context: str) -> None:
    """Print a clear message and exit on auth/config failures (avoid 1000s of silent flags)."""
    if resp.status_code in (401, 403):
        print(
            f"\nERROR: Places API (New) returned HTTP {resp.status_code} ({context}).\n"
            "  Common fix: Google Cloud Console → APIs & Services → Enable APIs →\n"
            '  enable "Places API (New)" for this project, and ensure billing is active if required.\n'
            "  This script uses: https://places.googleapis.com/v1/places:searchText\n\n"
            f"  Response body (first 800 chars):\n  {resp.text[:800]!r}\n",
            file=sys.stderr,
        )
        sys.exit(1)
    if resp.status_code >= 400:
        print(
            f"\nERROR: Places API (New) returned HTTP {resp.status_code} ({context}).\n"
            f"  Response body (first 800 chars):\n  {resp.text[:800]!r}\n",
            file=sys.stderr,
        )
        sys.exit(1)


# ── Step 1: Load data ─────────────────────────────────────────────────────────

print("Loading input file...")
df = pd.read_excel(INPUT_FILE, header=2)
df.columns = ["num", "club_name", "city"]
df = df.dropna(subset=["club_name"]).copy()
df["num"] = df["num"].astype(str)
print(f"  Loaded {len(df)} clubs")

# ── Step 2: Name-based pre-filter ─────────────────────────────────────────────

ELIMINATE_PATTERNS = [
    r"WATERPOLO", r"POLO\s", r"W\.P\.", r"\sWP\s", r"C\.\s+W\.",
    r"SINCRONIZADA", r"SINCRONI", r"SINCRO\b", r"SIN\.",
    r"ARTISTICA", r"ARTÍSTICA",
]
COUNTER_SIGNALS = [r"NATACION", r"NATACIÓ", r"C\.N\."]

elim_re  = re.compile("|".join(ELIMINATE_PATTERNS), re.IGNORECASE)
keep_re  = re.compile("|".join(COUNTER_SIGNALS),    re.IGNORECASE)

def classify_by_name(name: str) -> str:
    if elim_re.search(name) and not keep_re.search(name):
        return "eliminate"
    return "pending"

df["status"] = df["club_name"].apply(classify_by_name)
df["elimination_reason"] = df["status"].apply(
    lambda s: "name filter: waterpolo/artistica/sincronizada" if s == "eliminate" else ""
)

eliminated = (df["status"] == "eliminate").sum()
pending    = (df["status"] == "pending").sum()
print(f"  Name filter: {eliminated} eliminated, {pending} pending")

# ── Step 3: Places API enrichment ─────────────────────────────────────────────

new_cols = ["places_name", "formatted_address", "lat", "lng",
            "website", "phone", "opening_hours", "primary_type", "all_types"]
for col in new_cols:
    df[col] = ""

if not API_KEY:
    print("\nERROR: GOOGLE_PLACES_API_KEY not found in .env.local — skipping enrichment.")
else:
    print("\nProbing Places API (New) with a single request...")
    probe = places_search_text("Piscina Municipal Madrid")
    exit_if_places_api_blocked(probe, "startup probe")
    print("  Probe OK — Places API (New) is reachable.\n")

    pending_idx = df[df["status"] == "pending"].index
    total = len(pending_idx)
    print(f"Enriching {total} clubs via Places API...")

    for i, idx in enumerate(pending_idx, 1):
        club = df.at[idx, "club_name"]
        city = df.at[idx, "city"] or ""
        query = f"{club} {city} España"

        try:
            resp = places_search_text(query)
            # First real row: fail fast if key revoked mid-run (rare)
            if i == 1:
                exit_if_places_api_blocked(resp, "first club request")
            resp.raise_for_status()
            places = resp.json().get("places", [])

            if places:
                p = places[0]
                df.at[idx, "places_name"]      = p.get("displayName", {}).get("text", "")
                df.at[idx, "formatted_address"] = p.get("formattedAddress", "")
                loc = p.get("location", {})
                df.at[idx, "lat"]              = loc.get("latitude", "")
                df.at[idx, "lng"]              = loc.get("longitude", "")
                df.at[idx, "website"]          = p.get("websiteUri", "")
                df.at[idx, "phone"]            = p.get("nationalPhoneNumber", "")
                hours = p.get("regularOpeningHours", {}).get("weekdayDescriptions", [])
                df.at[idx, "opening_hours"]    = " | ".join(hours)
                df.at[idx, "primary_type"]     = p.get("primaryType", "")
                df.at[idx, "all_types"]        = ", ".join(p.get("types", []))
                df.at[idx, "status"] = "enriched"
                df.at[idx, "elimination_reason"] = ""
            else:
                df.at[idx, "status"] = "flag"
                df.at[idx, "elimination_reason"] = "no Places result found"

        except requests.HTTPError as e:
            df.at[idx, "status"] = "flag"
            df.at[idx, "elimination_reason"] = f"API error: {e}"
        except Exception as e:
            df.at[idx, "status"] = "flag"
            df.at[idx, "elimination_reason"] = f"API error: {e}"

        if i % 50 == 0 or i == total:
            n_enr = (df["status"] == "enriched").sum()
            n_flg = (df["status"] == "flag").sum()
            print(f"  {i}/{total} processed — enriched: {n_enr}, flagged: {n_flg}")

        time.sleep(REQUEST_DELAY)

# ── Step 4: Write output ──────────────────────────────────────────────────────

print(f"\nWriting output to {OUTPUT_FILE} ...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Final counts
total_clubs = len(df)
n_eliminated = (df["status"] == "eliminate").sum()
n_enriched   = (df["status"] == "enriched").sum()
n_flagged    = (df["status"] == "flag").sum()

wb = Workbook()
ws = wb.active
ws.title = "Enriched Clubs"

# Summary rows
header_font  = Font(name="Arial", bold=True, color="FFFFFF")
header_fill  = PatternFill("solid", start_color="1F3864")
summary_fill = PatternFill("solid", start_color="D9E1F2")

ws.append(["RFEN Swim Clubs — Enriched", "", "", "", "", "", "", "", "", "", "", "", ""])
ws["A1"].font = Font(name="Arial", bold=True, size=12)

ws.append([
    f"Total: {total_clubs}",
    f"Eliminated (name filter): {n_eliminated}",
    f"Enriched: {n_enriched}",
    f"Flagged (no result): {n_flagged}",
])
for cell in ws[2]:
    cell.font = Font(name="Arial", bold=True)
    cell.fill = summary_fill

ws.append([])  # blank row

# Column headers
columns = ["num", "club_name", "city", "status", "elimination_reason",
           "places_name", "formatted_address", "lat", "lng",
           "website", "phone", "opening_hours", "primary_type", "all_types"]
ws.append(columns)
for cell in ws[4]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Data rows
for _, row in df.iterrows():
    ws.append([str(row.get(c, "")) for c in columns])

# Column widths
widths = [6, 40, 20, 12, 35, 35, 40, 10, 10, 40, 18, 60, 20, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

wb.save(OUTPUT_FILE)
print(f"  Done. {OUTPUT_FILE}")
print(f"\nSummary:")
print(f"  Total clubs:        {total_clubs}")
print(f"  Eliminated:         {n_eliminated}")
print(f"  Enriched:           {n_enriched}")
print(f"  Flagged:            {n_flagged}")
